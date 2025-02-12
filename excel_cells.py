import openpyxl

wb = openpyxl.load_workbook('data/example.xlsx')
sheet = wb['Data']
print(sheet['A1'])
print(type(sheet['A1'].value))
print(sheet['A1'].value)
c = sheet['B1']
print(c)
print(type(c.value))
print(c.value)
print('Row %s, Column %s (%s) is %s' % (c.row, c.column, c.column_letter, c.value))
print(sheet.max_row)
print(sheet.max_column)

from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print(list(sheet.columns)[1]) # Get second column's cells.
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)