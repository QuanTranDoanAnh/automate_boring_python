import openpyxl
wb = openpyxl.Workbook() # Create a blank workbook.
print(wb.sheetnames) # It starts with one sheet.
sheet = wb.active
print(sheet.title) # The sheet's title is 'Sheet'.
sheet.title = 'Spam Bacon Eggs Sheet' # Change the title.
print(wb.sheetnames) # The title is updated.
sheet.title = 'Spam Spam Spam' # Change the title.
wb.save('data/example_copy.xlsx') # Save the workbook to a file.
wb.create_sheet(index=0, title='First Sheet') # Create a new sheet.

wb = openpyxl.Workbook() # Create a blank workbook.
print(wb.sheetnames) # It starts with one sheet.
wb.create_sheet() # Create a new sheet.
print(wb.sheetnames) # The new sheet is named 'Sheet1'.
wb.create_sheet(index=0, title='First Sheet') # Create a new sheet at index 0.
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)
wb.save('data/example_copy.xlsx') # Save the workbook to a file.

sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!' # Edit the cell's value.
print(sheet['A1'].value) # Print the cell's value.
wb.save('data/example_copy.xlsx') # Save the workbook to a file.